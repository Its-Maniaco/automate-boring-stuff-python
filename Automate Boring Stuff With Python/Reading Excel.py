import openpyxl
import os
os.chdir('D:\\Python Projects\\Automate Boring Stuff With Python')

workbook = openpyxl.load_workbook('example.xlsx')
#example.xlsx opened using load_workbook and stored into workbook object)
print(type(workbook))

# WAS -> workbook.get_sheet_names()
print(workbook.sheetnames)

# Was -> sheet = workbook.get_sheet_by_name('Sheet1')
sheet = workbook['Sheet1']
print(type(sheet))

cellO = sheet['A1'] # Gives a cell object
# or cellO = sheet.cell(row = 1, column = 1)
print(type(cellO))
cellO.value
print(cellO.value)
print(type(cellO.value))